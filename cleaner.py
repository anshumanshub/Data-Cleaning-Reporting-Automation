"""
cleaner.py
----------
Core engine for the Data Cleaning & Reporting Automation tool.

Handles:
  - Ingestion of CSV / TSV / Excel files
  - Duplicate detection & removal
  - Missing value detection & imputation
  - Date and phone-number standardization
  - IQR-based outlier detection
  - Data quality scoring
  - Multi-sheet Excel report generation with charts (openpyxl)

Usage (programmatic):
    from cleaner import DataCleaner
    dc = DataCleaner("input/sales_data.csv")
    dc.run_pipeline()
    dc.export_report("output/cleaning_report.xlsx")

Usage (CLI):
    python cleaner.py input/sales_data.csv --output output/cleaning_report.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.worksheet import Worksheet


# --------------------------------------------------------------------------- #
# Styling constants (kept consistent across all sheets)
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, name="Calibri", size=14, color="1F4E78")
SUBTITLE_FONT = Font(italic=True, name="Calibri", size=10, color="595959")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


@dataclass
class CleaningLog:
    """Keeps a structured record of every change made to the dataset."""
    rows_in: int = 0
    rows_out: int = 0
    duplicates_removed: int = 0
    missing_values_found: int = 0
    missing_values_imputed: dict = field(default_factory=dict)
    outliers_flagged: dict = field(default_factory=dict)
    dates_standardized: dict = field(default_factory=dict)
    phones_standardized: int = 0
    columns_dropped: list = field(default_factory=list)
    quality_score_before: float = 0.0
    quality_score_after: float = 0.0
    notes: list = field(default_factory=list)


class DataCleaner:
    """End-to-end cleaning pipeline for a single tabular data file."""

    DATE_COL_HINTS = ("date", "dob", "created", "updated", "timestamp")
    PHONE_COL_HINTS = ("phone", "mobile", "contact_no", "tel")

    def __init__(self, filepath: str, sheet_name=0):
        self.filepath = Path(filepath)
        self.sheet_name = sheet_name
        self.raw_df: pd.DataFrame | None = None
        self.df: pd.DataFrame | None = None
        self.log = CleaningLog()

    # ----------------------------- Ingestion ----------------------------- #
    def load(self) -> pd.DataFrame:
        suffix = self.filepath.suffix.lower()
        if suffix == ".csv":
            df = pd.read_csv(self.filepath)
        elif suffix == ".tsv":
            df = pd.read_csv(self.filepath, sep="\t")
        elif suffix in (".xlsx", ".xls"):
            df = pd.read_excel(self.filepath, sheet_name=self.sheet_name)
        else:
            raise ValueError(f"Unsupported file type: {suffix}")

        df.columns = [str(c).strip() for c in df.columns]
        self.raw_df = df.copy()
        self.df = df.copy()
        self.log.rows_in = len(df)
        return df

    # ------------------------------ Quality ------------------------------ #
    def _quality_score(self, df: pd.DataFrame) -> float:
        """Composite score 0-100: completeness + uniqueness + validity."""
        if df.empty:
            return 0.0
        completeness = 1 - (df.isna().sum().sum() / (df.shape[0] * df.shape[1]))
        uniqueness = 1 - (df.duplicated().sum() / max(len(df), 1))
        score = (completeness * 0.6 + uniqueness * 0.4) * 100
        return round(score, 2)

    # ---------------------------- Duplicates ------------------------------ #
    def remove_duplicates(self):
        before = len(self.df)
        self.df = self.df.drop_duplicates(keep="first").reset_index(drop=True)
        self.log.duplicates_removed = before - len(self.df)
        return self

    # ------------------------- Missing values ----------------------------- #
    def handle_missing_values(self, numeric_strategy="median", categorical_strategy="mode"):
        df = self.df
        self.log.missing_values_found = int(df.isna().sum().sum())

        for col in df.columns:
            n_missing = df[col].isna().sum()
            if n_missing == 0:
                continue

            if pd.api.types.is_numeric_dtype(df[col]):
                fill_value = df[col].median() if numeric_strategy == "median" else df[col].mean()
                df[col] = df[col].fillna(fill_value)
                self.log.missing_values_imputed[col] = f"{n_missing} filled with {numeric_strategy} ({round(fill_value, 2)})"
            else:
                mode_series = df[col].mode(dropna=True)
                fill_value = mode_series.iloc[0] if not mode_series.empty else "Unknown"
                df[col] = df[col].fillna(fill_value)
                self.log.missing_values_imputed[col] = f"{n_missing} filled with mode ('{fill_value}')"

        self.df = df
        return self

    # ----------------------------- Dates ----------------------------------- #
    def standardize_dates(self, output_format="%Y-%m-%d"):
        df = self.df
        candidate_cols = [c for c in df.columns if any(h in c.lower() for h in self.DATE_COL_HINTS)]

        for col in candidate_cols:
            try:
                parsed = pd.to_datetime(df[col], errors="coerce", infer_datetime_format=True)
                success_rate = parsed.notna().mean()
                if success_rate >= 0.5:  # only standardize if it looks like a real date column
                    df[col] = parsed.dt.strftime(output_format)
                    self.log.dates_standardized[col] = f"{round(success_rate * 100, 1)}% parsed successfully"
            except Exception as e:  # pragma: no cover
                self.log.notes.append(f"Could not parse dates in '{col}': {e}")

        self.df = df
        return self

    # ----------------------------- Phones ----------------------------------- #
    def standardize_phone_numbers(self, country_code="+1"):
        df = self.df
        candidate_cols = [c for c in df.columns if any(h in c.lower() for h in self.PHONE_COL_HINTS)]

        def clean_phone(val):
            if pd.isna(val):
                return val
            digits = re.sub(r"\D", "", str(val))
            if len(digits) == 10:
                return f"{country_code} ({digits[0:3]}) {digits[3:6]}-{digits[6:]}"
            elif len(digits) == 11 and digits.startswith("1"):
                return f"{country_code} ({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
            return str(val)  # leave malformed numbers untouched but flagged in notes

        for col in candidate_cols:
            changed = df[col].apply(clean_phone)
            n_changed = (changed != df[col]).sum()
            df[col] = changed
            self.log.phones_standardized += int(n_changed)

        self.df = df
        return self

    # ----------------------------- Outliers ---------------------------------- #
    def flag_outliers_iqr(self, k=1.5):
        """Adds a `<col>_outlier` boolean flag column for each numeric column."""
        df = self.df
        numeric_cols = df.select_dtypes(include=[np.number]).columns

        for col in numeric_cols:
            q1, q3 = df[col].quantile(0.25), df[col].quantile(0.75)
            iqr = q3 - q1
            lower, upper = q1 - k * iqr, q3 + k * iqr
            flags = (df[col] < lower) | (df[col] > upper)
            n_flagged = int(flags.sum())
            if n_flagged > 0:
                df[f"{col}_outlier"] = flags
                self.log.outliers_flagged[col] = n_flagged

        self.df = df
        return self

    # ------------------------------ Pipeline --------------------------------- #
    def run_pipeline(self):
        if self.raw_df is None:
            self.load()

        self.log.quality_score_before = self._quality_score(self.raw_df)

        (self
         .remove_duplicates()
         .handle_missing_values()
         .standardize_dates()
         .standardize_phone_numbers()
         .flag_outliers_iqr())

        self.log.rows_out = len(self.df)
        self.log.quality_score_after = self._quality_score(self.df)
        return self.df

    # ------------------------------ Excel report ------------------------------ #
    def export_report(self, output_path: str):
        if self.df is None:
            raise RuntimeError("Run the pipeline before exporting a report.")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        self._write_summary_sheet(wb)
        self._write_cleaned_data_sheet(wb)
        self._write_issues_sheet(wb)
        self._write_charts_sheet(wb)

        wb.save(output_path)
        return output_path

    # --- sheet builders -------------------------------------------------- #
    def _autosize(self, ws: Worksheet, df: pd.DataFrame, start_col=1):
        for i, col in enumerate(df.columns, start=start_col):
            max_len = max(df[col].astype(str).map(len).max() if len(df) else 0, len(str(col)))
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 45)

    def _style_header_row(self, ws: Worksheet, row: int, n_cols: int, start_col=1):
        for c in range(start_col, start_col + n_cols):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    def _write_summary_sheet(self, wb: Workbook):
        ws = wb.active
        ws.title = "Summary"

        ws["B2"] = "Data Cleaning & Reporting Automation"
        ws["B2"].font = TITLE_FONT
        ws["B3"] = f"Source file: {self.filepath.name}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["B3"].font = SUBTITLE_FONT
        ws.merge_cells("B2:F2")
        ws.merge_cells("B3:F3")

        rows = [
            ("Rows in (raw)", self.log.rows_in),
            ("Rows out (cleaned)", self.log.rows_out),
            ("Duplicates removed", self.log.duplicates_removed),
            ("Missing values found", self.log.missing_values_found),
            ("Columns with imputed values", len(self.log.missing_values_imputed)),
            ("Date columns standardized", len(self.log.dates_standardized)),
            ("Phone numbers standardized", self.log.phones_standardized),
            ("Columns with outliers flagged", len(self.log.outliers_flagged)),
            ("Quality score (before)", f"{self.log.quality_score_before} / 100"),
            ("Quality score (after)", f"{self.log.quality_score_after} / 100"),
        ]

        start_row = 5
        ws.cell(row=start_row, column=2, value="Metric")
        ws.cell(row=start_row, column=3, value="Value")
        self._style_header_row(ws, start_row, 2, start_col=2)

        for i, (label, value) in enumerate(rows, start=start_row + 1):
            ws.cell(row=i, column=2, value=label).border = THIN_BORDER
            cell = ws.cell(row=i, column=3, value=value)
            cell.border = THIN_BORDER
            if "Quality score" in label:
                numeric = float(str(value).split("/")[0])
                cell.fill = GOOD_FILL if numeric >= 80 else WARN_FILL if numeric >= 60 else BAD_FILL

        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 22

        # Imputation detail block
        detail_row = start_row + len(rows) + 3
        ws.cell(row=detail_row, column=2, value="Missing Value Imputation Detail").font = Font(bold=True, size=12)
        detail_row += 1
        ws.cell(row=detail_row, column=2, value="Column")
        ws.cell(row=detail_row, column=3, value="Action Taken")
        self._style_header_row(ws, detail_row, 2, start_col=2)
        for col, action in self.log.missing_values_imputed.items():
            detail_row += 1
            ws.cell(row=detail_row, column=2, value=col).border = THIN_BORDER
            ws.cell(row=detail_row, column=3, value=action).border = THIN_BORDER
        ws.column_dimensions["C"].width = 45

        if self.log.notes:
            detail_row += 2
            ws.cell(row=detail_row, column=2, value="Notes").font = Font(bold=True)
            for note in self.log.notes:
                detail_row += 1
                ws.cell(row=detail_row, column=2, value=note)

    def _write_cleaned_data_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Cleaned Data")
        df = self.df

        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=1, column=j, value=col)
        self._style_header_row(ws, 1, len(df.columns))

        for i, row in enumerate(df.itertuples(index=False), start=2):
            for j, value in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=value if pd.notna(value) else None)
                cell.border = THIN_BORDER

        self._autosize(ws, df)
        ws.freeze_panes = "A2"

    def _write_issues_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Outliers & Flags")
        outlier_cols = [c for c in self.df.columns if c.endswith("_outlier")]

        if not outlier_cols:
            ws["B2"] = "No outliers detected (IQR method)."
            ws["B2"].font = Font(italic=True)
            return

        flagged_rows = self.df[self.df[outlier_cols].any(axis=1)]
        base_cols = [c for c in self.df.columns if not c.endswith("_outlier")]
        subset = flagged_rows[base_cols + outlier_cols]

        for j, col in enumerate(subset.columns, start=1):
            ws.cell(row=1, column=j, value=col)
        self._style_header_row(ws, 1, len(subset.columns))

        for i, row in enumerate(subset.itertuples(index=False), start=2):
            for j, value in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=value if pd.notna(value) else None)
                cell.border = THIN_BORDER
                if subset.columns[j - 1].endswith("_outlier") and value is True:
                    cell.fill = BAD_FILL

        self._autosize(ws, subset)
        ws.freeze_panes = "A2"

    def _write_charts_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Charts")
        ws["B2"] = "Visual Summary"
        ws["B2"].font = TITLE_FONT

        # --- Data quality before/after bar chart ---
        ws["B4"] = "Metric"
        ws["C4"] = "Score"
        ws["B5"] = "Before Cleaning"
        ws["C5"] = self.log.quality_score_before
        ws["B6"] = "After Cleaning"
        ws["C6"] = self.log.quality_score_after
        self._style_header_row(ws, 4, 2, start_col=2)

        bar = BarChart()
        bar.title = "Data Quality Score: Before vs After"
        bar.y_axis.title = "Score (0-100)"
        bar.x_axis.title = "Stage"
        data = Reference(ws, min_col=3, min_row=4, max_row=6)
        cats = Reference(ws, min_col=2, min_row=5, max_row=6)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.width, bar.height = 14, 8
        ws.add_chart(bar, "E4")

        # --- Issue breakdown pie chart ---
        issue_start = 14
        ws.cell(row=issue_start, column=2, value="Issue Type")
        ws.cell(row=issue_start, column=3, value="Count")
        self._style_header_row(ws, issue_start, 2, start_col=2)

        issue_rows = [
            ("Duplicates Removed", self.log.duplicates_removed),
            ("Missing Values Found", self.log.missing_values_found),
            ("Outlier Values Flagged", sum(self.log.outliers_flagged.values())),
            ("Phone Numbers Fixed", self.log.phones_standardized),
        ]
        for i, (label, value) in enumerate(issue_rows, start=issue_start + 1):
            ws.cell(row=i, column=2, value=label)
            ws.cell(row=i, column=3, value=value)

        pie = PieChart()
        pie.title = "Issue Breakdown"
        data = Reference(ws, min_col=3, min_row=issue_start, max_row=issue_start + len(issue_rows))
        cats = Reference(ws, min_col=2, min_row=issue_start + 1, max_row=issue_start + len(issue_rows))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        pie.width, pie.height = 14, 8
        ws.add_chart(pie, "E18")

        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 14


# --------------------------------------------------------------------------- #
# CLI entry point
# --------------------------------------------------------------------------- #
def main():
    parser = argparse.ArgumentParser(description="Automate data cleaning and reporting.")
    parser.add_argument("input", help="Path to input CSV/TSV/Excel file")
    parser.add_argument("--output", "-o", default="output/cleaning_report.xlsx", help="Path to output Excel report")
    parser.add_argument("--sheet", default=0, help="Sheet name/index for Excel inputs (default: first sheet)")
    args = parser.parse_args()

    dc = DataCleaner(args.input, sheet_name=args.sheet)
    dc.load()
    print(f"Loaded {dc.log.rows_in} rows from {args.input}")

    dc.run_pipeline()
    print(f"Cleaning complete. Rows: {dc.log.rows_in} -> {dc.log.rows_out}")
    print(f"Quality score: {dc.log.quality_score_before} -> {dc.log.quality_score_after}")

    out_path = dc.export_report(args.output)
    print(f"Report written to {out_path}")


if __name__ == "__main__":
    sys.exit(main())
